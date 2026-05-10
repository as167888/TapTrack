import db
import openpyxl
import os
from datetime import datetime

PAGE_SIZE = 20
OUTPUT_DIR = "./export/view"


def _fmt_num(val):
    """Format a number string with commas for display."""
    if val is None or val == "" or val == "获取失败":
        return "-"
    try:
        n = int(str(val).replace(",", ""))
        if n >= 100000000:
            return f"{n/100000000:.2f}亿"
        elif n >= 10000:
            return f"{n/10000:.1f}万"
        else:
            return f"{n:,}"
    except (ValueError, TypeError):
        return str(val)


def _print_header(title):
    print()
    print("=" * 60)
    print(f"  {title}")
    print("=" * 60)


def show_stats():
    _print_header("数据库统计")
    total = db.get_stats()
    print(f"  游戏总数: {total}")
    print()
    print("  分类统计:")
    print(f"  {'分类':<16} {'数量':>6}")
    print(f"  {'-'*16} {'-'*6}")
    for cat, count in db.get_game_count_by_category():
        print(f"  {cat:<16} {count:>6}")
    print()
    input("按回车键返回...")


def browse_all():
    offset = 0
    while True:
        rows, total = db.get_paginated_games(offset, PAGE_SIZE)
        if not rows:
            print("\n暂无数据")
            input("按回车键返回...")
            return

        _print_header(f"浏览游戏列表 (第 {offset + 1}-{min(offset + PAGE_SIZE, total)} 条, 共 {total} 条)")
        print(f"  {'ID':<6} {'名称':<20} {'分类':<10} {'添加时间':<20} {'详情页链接'}")
        print(f"  {'-'*6} {'-'*20} {'-'*10} {'-'*20} {'-'*56}")
        for row in rows:
            gid, name, app_id, url, cat, created = row
            display_name = name[:18] + ".." if len(name) > 20 else name
            display_cat = (cat or "-")[:8]
            display_url = url[:53] + "..." if len(url) > 56 else url
            print(f"  {gid:<6} {display_name:<20} {display_cat:<10} {created:<20} {display_url}")

        print()
        print(f"  [N] 下一页  [P] 上一页  [E] 导出当前页  [Q] 返回")
        choice = input("  请选择: ").strip().upper()

        if choice == "N":
            if offset + PAGE_SIZE < total:
                offset += PAGE_SIZE
            else:
                print("  已经是最后一页")
        elif choice == "P":
            if offset >= PAGE_SIZE:
                offset -= PAGE_SIZE
            else:
                print("  已经是第一页")
        elif choice == "E":
            _export_to_excel(rows)
        elif choice == "Q":
            return


def browse_by_category():
    cats = db.get_all_categories()
    if not cats:
        print("\n暂无分类数据")
        input("按回车键返回...")
        return

    _print_header("按分类筛选")
    for i, cat in enumerate(cats, 1):
        print(f"  {i}. {cat}")

    print()
    choice = input("请选择分类编号 (0 返回): ").strip()
    if choice == "0" or not choice:
        return

    try:
        idx = int(choice) - 1
        if idx < 0 or idx >= len(cats):
            print("无效选择")
            return
    except ValueError:
        print("无效输入")
        return

    selected_cat = cats[idx]
    rows = db.get_games_by_category(selected_cat)

    _print_header(f"分类: {selected_cat} ({len(rows)} 条)")
    print(f"  {'ID':<6} {'名称':<20} {'添加时间':<20} {'详情页链接'}")
    print(f"  {'-'*6} {'-'*20} {'-'*20} {'-'*56}")
    for row in rows:
        gid, name, app_id, url, cat, created = row
        display_name = name[:18] + ".." if len(name) > 20 else name
        display_url = url[:53] + "..." if len(url) > 56 else url
        print(f"  {gid:<6} {display_name:<20} {created:<20} {display_url}")

    print()
    print(f"  [E] 导出  [Q] 返回")
    choice = input("  请选择: ").strip().upper()
    if choice == "E":
        _export_to_excel(rows)
    elif choice == "Q":
        return


def search():
    _print_header("搜索游戏")
    keyword = input("请输入游戏名称关键词 (留空返回): ").strip()
    if not keyword:
        return

    rows = db.search_games(keyword)
    if not rows:
        print(f"\n未找到包含 \"{keyword}\" 的游戏")
        input("按回车键返回...")
        return

    _print_header(f"搜索结果: \"{keyword}\" ({len(rows)} 条)")
    print(f"  {'ID':<6} {'名称':<24} {'分类':<12} {'添加时间':<20}")
    print(f"  {'-'*6} {'-'*24} {'-'*12} {'-'*20}")
    for row in rows:
        gid, name, app_id, url, cat, created = row
        display_name = name[:22] + ".." if len(name) > 24 else name
        display_cat = (cat or "-")[:10]
        print(f"  {gid:<6} {display_name:<24} {display_cat:<12} {created:<20}")

    print()
    print(f"  [E] 导出  [D] 删除游戏  [Q] 返回")
    choice = input("  请选择: ").strip().upper()
    if choice == "E":
        _export_to_excel(rows)
    elif choice == "D":
        del_id = input("  输入要删除的游戏 ID: ").strip()
        try:
            if db.delete_game(int(del_id)):
                print(f"  已删除 ID={del_id}")
            else:
                print(f"  未找到 ID={del_id}")
        except ValueError:
            print("  无效 ID")
        input("按回车键继续...")


def _export_to_excel(rows):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(OUTPUT_DIR, f"{timestamp}_view_export.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "游戏数据"
    ws["A1"] = "ID"
    ws["B1"] = "游戏名称"
    ws["C1"] = "App ID"
    ws["D1"] = "详情页链接"
    ws["E1"] = "分类"
    ws["F1"] = "添加时间"
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 22

    for i, row in enumerate(rows, start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)

    wb.save(filepath)
    print(f"  已导出: {filepath}")


# ---- 爬取详情浏览 ----

def _show_detail_summary():
    """显示爬取详情数据库概要统计。"""
    _print_header("爬取详情 — 数据库概要")
    summary = db.get_detail_summary()

    print(f"  爬取记录总数: {summary['total_records']}")
    print(f"  覆盖游戏数:   {summary['distinct_games']}")
    print(f"  爬取会话数:   {summary['session_count']}")

    if summary["latest_session"]:
        sid, sat = summary["latest_session"]
        print(f"  最近爬取:     {sat} (session #{sid}, {summary['latest_count']} 条记录)")
    else:
        print("  最近爬取:     暂无")

    # Top 10 下载量
    print()
    print("  —— 下载量 Top 10 (最近一次爬取) ——")
    print(f"  {'排名':<4} {'游戏名称':<22} {'下载量':<12} {'评分':<6} {'关注量':<10}")
    print(f"  {'-'*4} {'-'*22} {'-'*12} {'-'*6} {'-'*10}")
    top_dl = db.get_top_by_downloads(10)
    if top_dl:
        for i, (name, dl, rating, followers, _) in enumerate(top_dl, 1):
            display_name = name[:20] + ".." if len(name) > 22 else name
            print(f"  {i:<4} {display_name:<22} {_fmt_num(dl):<12} {rating or '-':<6} {_fmt_num(followers):<10}")
    else:
        print("  (暂无数据)")

    # Top 10 评分
    print()
    print("  —— 评分 Top 10 (最近一次爬取) ——")
    print(f"  {'排名':<4} {'游戏名称':<22} {'评分':<6} {'下载量':<12}")
    print(f"  {'-'*4} {'-'*22} {'-'*6} {'-'*12}")
    top_rt = db.get_top_by_rating(10)
    if top_rt:
        for i, (name, rating, dl, _, _) in enumerate(top_rt, 1):
            display_name = name[:20] + ".." if len(name) > 22 else name
            print(f"  {i:<4} {display_name:<22} {rating or '-':<6} {_fmt_num(dl):<12}")
    else:
        print("  (暂无数据)")

    input("\n按回车键继续浏览明细...")


def _browse_session_records(session_id):
    """分页浏览某个 session 的爬取记录。"""
    offset = 0
    total = db.get_crawl_records_count(session_id)
    while True:
        rows = db.get_crawl_records(session_id, offset, PAGE_SIZE)
        if not rows:
            print("\n暂无数据")
            input("按回车键返回...")
            return

        _print_header(f"Session #{session_id} 爬取记录 (第 {offset + 1}-{min(offset + PAGE_SIZE, total)} 条, 共 {total} 条)")
        print(f"  {'ID':<5} {'游戏名称':<22} {'下载量':<12} {'评分':<6} {'关注量':<10} {'发布日期':<12}")
        print(f"  {'-'*5} {'-'*22} {'-'*12} {'-'*6} {'-'*10} {'-'*12}")
        for row in rows:
            rid, name, url, pub_date, dl, followers, rating, rc, crawled_at = row
            display_name = name[:20] + ".." if len(name) > 22 else name
            print(f"  {rid:<5} {display_name:<22} {_fmt_num(dl):<12} {rating or '-':<6} {_fmt_num(followers):<10} {pub_date or '-':<12}")

        print()
        print(f"  [N] 下一页  [P] 上一页  [Q] 返回")
        choice = input("  请选择: ").strip().upper()
        if choice == "N":
            if offset + PAGE_SIZE < total:
                offset += PAGE_SIZE
            else:
                print("  已经是最后一页")
        elif choice == "P":
            if offset >= PAGE_SIZE:
                offset -= PAGE_SIZE
            else:
                print("  已经是第一页")
        elif choice == "Q":
            return


def browse_details():
    """爬取详情浏览入口：先显示概要，再选择 session 浏览明细。"""
    db.init_db()
    db.init_detail_tables()

    sessions = db.get_crawl_sessions()
    if not sessions:
        print('\n暂无爬取详情数据。请先执行"爬取详情"(主菜单选项 3) 生成数据。')
        input("按回车键返回...")
        return

    # 先显示概要
    _show_detail_summary()

    # 再选择 session 浏览明细
    while True:
        _print_header("选择爬取会话")
        for sid, sat, cnt in sessions:
            print(f"  {sid}. {sat} — {cnt} 条记录")

        print()
        print("  输入 session 编号浏览明细，输入 0 返回")

        choice = input("请选择: ").strip()
        if choice == "0" or not choice:
            return

        try:
            sid = int(choice)
            valid_ids = [s[0] for s in sessions]
            if sid not in valid_ids:
                print("无效的 session 编号")
                continue
        except ValueError:
            print("无效输入")
            continue

        _browse_session_records(sid)


def run():
    db.init_db()
    db.init_detail_tables()

    while True:
        _print_header("数据库查看器")
        total = db.get_stats()
        print(f"  当前数据库共有 {total} 条游戏记录\n")
        print("  1. 浏览全部游戏（分页）")
        print("  2. 按分类筛选")
        print("  3. 搜索游戏")
        print("  4. 爬取详情 - 浏览历史爬取的详情数据")
        print("  0. 返回主菜单")
        print()

        choice = input("请选择: ").strip()

        if choice == "1":
            browse_all()
        elif choice == "2":
            browse_by_category()
        elif choice == "3":
            search()
        elif choice == "4":
            browse_details()
        elif choice == "0":
            break
        else:
            print("无效选项，请重试。")


if __name__ == "__main__":
    run()
