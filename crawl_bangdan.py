import requests
from bs4 import BeautifulSoup
import openpyxl
import os
import re
import time
from datetime import datetime
import db

LINKS_FILE = "榜单链接.txt"
OUTPUT_DIR = "./export/bangdan"
BASE_URL = "https://www.taptap.cn"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/125.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
}


def parse_links(filepath):
    """Extract (category_name, url) pairs from the links file."""
    pairs = []
    current_category = ""
    with open(filepath, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            parts = line.split("\t")
            if len(parts) >= 2:
                cat = parts[0].strip()
                url = parts[1].strip() if len(parts) >= 2 else ""
                if url.startswith("http"):
                    current_category = cat
                    pairs.append((current_category, url))
                elif cat and not url:
                    current_category = cat
    return pairs


def extract_games(html):
    """Extract game name and /app/XXXX link from page HTML."""
    soup = BeautifulSoup(html, "lxml")
    games = []

    for item in soup.find_all("div", class_="list-item"):
        name_meta = item.find("meta", itemprop="name")
        if not name_meta:
            continue
        name = (name_meta.get("content") or "").strip()
        if not name:
            continue

        app_link = None
        for a in item.find_all("a", href=True):
            href = a["href"]
            if re.match(r"^/app/\d+", href):
                app_link = href
                break

        if app_link:
            # Extract app ID from /app/XXXXX
            app_id = re.search(r"/app/(\d+)", app_link).group(1)
            games.append((name, app_id, app_link))

    return games


def run():
    # Initialize database
    db.init_db()

    pairs = parse_links(LINKS_FILE)
    print(f"共读取到 {len(pairs)} 个榜单链接")

    new_games = []
    new_count = 0
    exist_count = 0

    for idx, (category, url) in enumerate(pairs, 1):
        try:
            print(f"[{idx}/{len(pairs)}] 抓取 [{category}]: {url}")
            resp = requests.get(url, headers=HEADERS, timeout=30)
            resp.encoding = "utf-8"

            if resp.status_code != 200:
                print(f"  -> HTTP {resp.status_code}, 跳过")
                continue

            games = extract_games(resp.text)
            page_new = 0
            page_exist = 0
            for name, app_id, app_link in games:
                full_url = f"{BASE_URL}{app_link}?os=android"
                inserted = db.insert_game(name, app_id, full_url, category)
                if inserted:
                    new_count += 1
                    page_new += 1
                    new_games.append((name, full_url))
                else:
                    exist_count += 1
                    page_exist += 1

            print(f"  -> {len(games)} 个游戏 (新增 {page_new}, 已存在 {page_exist})")

        except Exception as e:
            print(f"  -> 出错: {e}")

        time.sleep(0.5)

    # Better summary after processing
    print(f"\n本次: 新增 {new_count} 个, 已存在 {exist_count} 个")

    # Get all games from DB
    all_games = db.get_all_games()

    # Export to Excel
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{timestamp}_bangdan_games.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)

    wb = openpyxl.Workbook()

    # Sheet 1: New games from this run
    ws_new = wb.active
    ws_new.title = "本次新增"
    ws_new["A1"] = "游戏名称"
    ws_new["B1"] = "详情页链接"
    ws_new.column_dimensions["A"].width = 30
    ws_new.column_dimensions["B"].width = 60
    for i, (name, link) in enumerate(new_games, start=2):
        ws_new.cell(row=i, column=1, value=name)
        ws_new.cell(row=i, column=2, value=link)

    # Sheet 2: All games in database
    ws_all = wb.create_sheet("数据库全量")
    ws_all["A1"] = "游戏名称"
    ws_all["B1"] = "详情页链接"
    ws_all.column_dimensions["A"].width = 30
    ws_all.column_dimensions["B"].width = 60
    for i, (name, link) in enumerate(all_games, start=2):
        ws_all.cell(row=i, column=1, value=name)
        ws_all.cell(row=i, column=2, value=link)

    wb.save(filepath)

    total = db.get_stats()
    print(f"数据库总计: {total} 个游戏")
    print(f"输出文件: {filepath}")


if __name__ == "__main__":
    run()
